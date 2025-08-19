import os
from cryptography.fernet import Fernet
import base64
import datetime
from datetime import datetime as dt
import openpyxl
from openpyxl import load_workbook
import json
import sys
import customtkinter as ctk
import datetime
import hashlib
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

LICENSE_KEY_FILE = "license.key"       # used for storing the encrypted license key
LICENSE_META_FILE = "license.json"     # used for storing expiry date etc.


SECRET_KEY = b'a_WeqpQk65a-PGnLFodDaeL7PnbRUDKv0VXwRr-sKKI='
fernet = Fernet(SECRET_KEY)
# you choose this value
def validate_license_key(key):
    try:
        # decrypt the license key
        data = fernet.decrypt(key.encode()).decode()

        # check if it's the admin master key
        if data == "ADMIN":
            return True, "Admin license (no expiry)"

        # otherwise, treat it as an expiry date
        exp_date = dt.strptime(data, "%Y-%m-%d")
        if exp_date < dt.now():

            return False, "License expired"
        return True, "License valid"
    except Exception:
        return False, "Invalid license key"


def get_license_gui():
    # Check if license already exists
    if os.path.exists(LICENSE_KEY_FILE):
        with open(LICENSE_KEY_FILE, "r") as f:
            key = f.read().strip()
            valid, msg = validate_license_key(key)
            if valid:
                return True
            else:
                messagebox.showerror("License Error", msg)
                return False

    # Otherwise, ask user for license via GUI
    root = ctk.CTk()
    root.title("License Activation")
    root.geometry("400x200")
    root.resizable(False, False)

    label = ctk.CTkLabel(root, text="Enter your license key:", font=("Arial", 14))
    label.pack(pady=20)

    entry = ctk.CTkEntry(root, width=300)
    entry.pack(pady=10)

    result = {"valid": False}

    def submit_key():
        key = entry.get().strip()
        valid, msg = validate_license_key(key)
        if valid:
            with open(LICENSE_KEY_FILE, "w") as f:
                f.write(key)
            messagebox.showinfo("Success", "License activated successfully!")
            result["valid"] = True
            root.destroy()
        else:
            messagebox.showerror("Invalid License", msg)

    btn = ctk.CTkButton(root, text="Activate", command=submit_key)
    btn.pack(pady=15)

    root.mainloop()
    return result["valid"]


if not get_license_gui():
    sys.exit(1)


# always use the folder where the exe is located
base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
file_path = os.path.join(base_path, "MonaGraph.xlsm")

try:
    wb = load_workbook(file_path)
except Exception as e:
    print(f"Failed to open workbook: {e}")  # debug check

# right after you define file_path
if not os.path.exists(file_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"
    ws.append(["ID", "Name", "Age", "Diagnosis"])  # <-- adjust headers if needed
    wb.save(file_path)
else:
    wb = openpyxl.load_workbook(file_path)


# Load the workbook


APP_TITLE = "MonaGraph App"
USERS_FILE = "users.json"
CONFIG_FILE = "config.json"
LICENSE_META_FILE = "license.json"
SALT = "tdm_salt_2025"

# ---------- Utilities ----------
def hash_pw(pw: str) -> str:
    return hashlib.sha256((SALT + pw).encode("utf-8")).hexdigest()

def add_months(d: datetime.date, months: int) -> datetime.date:
    """Add calendar months to a date (no external libs)."""
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    # days in month
    dim = [31, 29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28,
           31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month-1]
    day = min(d.day, dim)
    return datetime.date(year, month, day)

# ---------- Bootstrap files if missing ----------
def ensure_default_users():
    if not os.path.exists(USERS_FILE):
        users = [
            {"username": "admin",  "role": "admin", "password_hash": hash_pw("admin123")},
            {"username": "user1",  "role": "user",  "password_hash": hash_pw("user1pass")},
            {"username": "user2", "role": "user",  "password_hash": hash_pw("user2pass")},
            {"username": "user3", "role": "user",  "password_hash": hash_pw("user3pass")},
        ]
        with open(USERS_FILE, "w") as f:
            json.dump(users, f, indent=2)

def ensure_default_config():
    if not os.path.exists(CONFIG_FILE):
        cfg = {"workbook_file": "MonaGraph.xlsm", "months_valid": 12}
        with open(CONFIG_FILE, "w") as f:
            json.dump(cfg, f, indent=2)

def load_config():
    ensure_default_config()
    with open(CONFIG_FILE, "r") as f:
        return json.load(f)

# ---------- License ----------
from typing import Tuple

def check_or_init_license(months_valid: int) -> Tuple[bool, str]:
    today = datetime.date.today()
    if not os.path.exists(LICENSE_META_FILE):
        data = {"first_use": today.isoformat()}
        with open(LICENSE_META_FILE, "w") as f:
            json.dump(data, f, indent=2)
        expiry = add_months(today, months_valid)
        return True, f"License created. Expires on {expiry}."
    with open(LICENSE_META_FILE, "r") as f:
        data = json.load(f)
    first_use = datetime.date.fromisoformat(data["first_use"])
    expiry = add_months(first_use, months_valid)
    if today > expiry:
        return False, f"Your license expired on {expiry}."
    return True, f"License OK. Expires on {expiry}."

def admin_renew_license(months_valid: int):
    today = datetime.date.today()
    with open(LICENSE_META_FILE, "w") as f:
        json.dump({"first_use": today.isoformat()}, f, indent=2)
    messagebox.showinfo("License", f"License renewed. New expiry: {add_months(today, months_valid)}")

# ---------- Users ----------
def load_users():
    ensure_default_users()
    with open(USERS_FILE, "r") as f:
        return json.load(f)

def verify_login(username: str, password: str):
    h = hash_pw(password)
    for u in load_users():
        if u["username"] == username and u["password_hash"] == h:
            return u
    return None

# ---------- Excel helpers ----------
def get_patient_sheets(wb):
    # CHANGED: make detection case-insensitive and strip spaces
    sheets = []
    for ws in wb.worksheets:
        name = ws.title.strip()
        if name.lower().startswith("patient"):
            sheets.append(name)
    return sheets

def next_empty_row(ws, start_row=15):
    """First row from start_row where B:G are all empty (we’ll still use the first empty row to write extended columns)."""
    r = start_row
    while True:
        vals = [ws.cell(row=r, column=c).value for c in range(2, 8)]
        if all(v is None or str(v).strip() == "" for v in vals):
            return r
        r += 1

def write_patient_details(ws, details, force_update=False):
    # B2..B6 labeled values
    def empty_or_missing(cell):
        v = ws[cell].value
        return v is None or str(v).strip() == ""

    if force_update or empty_or_missing("B2"):
        ws["B2"] = f"Patient Name: {details.get('name','')}"
    if force_update or empty_or_missing("B3"):
        ws["B3"] = f"Sex: {details.get('sex','')}"
    if force_update or empty_or_missing("B4"):
        ws["B4"] = f"Genotype: {details.get('genotype','')}"
    if force_update or empty_or_missing("B5"):
        ws["B5"] = f"Date of Birth: {details.get('dob','')}"
    if force_update or empty_or_missing("B6"):
        ws["B6"] = f"Contact Number: {details.get('contact','')}"

def append_daily_record(ws, rec):
    """
    Writes daily record to:
    B Date, C Dose, D Weight, E BMI, F INR, G Notes,
    H Blood Glucose, I Systolic BP, J Diastolic BP, K Height, L Lung Capacity
    Only writes keys present in rec (so user can choose variables).
    """
    r = next_empty_row(ws, 15)

    # Original columns kept exactly
    if "date" in rec:   ws.cell(row=r, column=2).value = rec.get("date", "")       # B
    if "dose" in rec:   ws.cell(row=r, column=3).value = rec.get("dose", "")       # C
    if "weight" in rec: ws.cell(row=r, column=4).value = rec.get("weight", "")     # D
    if "bmi" in rec:    ws.cell(row=r, column=5).value = rec.get("bmi", "")        # E
    if "inr" in rec:    ws.cell(row=r, column=6).value = rec.get("inr", "")        # F
    if "notes" in rec:  ws.cell(row=r, column=7).value = rec.get("notes", "")      # G

    # New variables to the right
    if "blood_glucose" in rec:  ws.cell(row=r, column=8).value  = rec.get("blood_glucose", "")  # H
    if "sbp" in rec:            ws.cell(row=r, column=9).value  = rec.get("sbp", "")            # I
    if "dbp" in rec:            ws.cell(row=r, column=10).value = rec.get("dbp", "")            # J
    if "height" in rec:         ws.cell(row=r, column=11).value = rec.get("height", "")         # K
    if "lung_capacity" in rec:  ws.cell(row=r, column=12).value = rec.get("lung_capacity", "")  # L
    return r

# ---------- Dark theme helpers ----------
def apply_dark_theme(root: tk.Tk):
    root.configure(bg="black")

    # Tk default options
    root.option_add("*Background", "black")
    root.option_add("*foreground", "white")
    root.option_add("*selectBackground", "#444444")
    root.option_add("*selectForeground", "white")
    root.option_add("*Entry.Background", "#222222")
    root.option_add("*Entry.Foreground", "white")
    root.option_add("*Entry.insertBackground", "white")
    root.option_add("*Text.Background", "#222222")
    root.option_add("*Text.Foreground", "white")
    root.option_add("*Listbox.Background", "#222222")
    root.option_add("*Listbox.Foreground", "white")
    root.option_add("*Button.Background", "#333333")
    root.option_add("*Button.Foreground", "white")

    # ttk theme
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass

    style.configure(".", background="black", foreground="white")
    style.configure("TLabel", background="black", foreground="white")
    style.configure("TFrame", background="black", foreground="white")
    style.configure("TLabelframe", background="black", foreground="white")
    style.configure("TLabelframe.Label", background="black", foreground="white")
    style.configure("TButton", background="#333333", foreground="white")
    style.map("TButton", background=[("active", "#444444")])
    style.configure("TCheckbutton", background="black", foreground="white")
    style.configure("TCombobox",
                    fieldbackground="#222222",
                    background="#333333",
                    foreground="white")
    style.map("TCombobox",
              fieldbackground=[("readonly", "#222222")],
              foreground=[("readonly", "white")])

def make_dark(frame: tk.Widget):
    # recursively set bg/fg for Tk (not ttk) widgets created explicitly
    try:
        frame.configure(bg="black")
    except Exception:
        pass
    for child in frame.winfo_children():
        cls = child.winfo_class().lower()
        if cls in ("frame", "labelframe"):
            make_dark(child)
        else:
            try:
                if hasattr(child, "configure"):
                    # set where supported
                    if isinstance(child, (tk.Label, tk.LabelFrame)):
                        child.configure(bg="black", fg="white")
                    elif isinstance(child, (tk.Entry, tk.Text)):
                        child.configure(bg="#222222", fg="white", insertbackground="white")
                    elif isinstance(child, tk.Button):
                        child.configure(bg="#333333", fg="white", activebackground="#444444")
            except Exception:
                pass

# ---------- App UI ----------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("MonaGraph")
        footer = tk.Label(
        self,
         text="MonaGraph® is an innovation by Pharmakon Axis",
         font=("Arial", 13, "italic"),
         fg="grey"
    )
        footer.pack(side="bottom", pady=5)
        self.geometry("1000x700")
        self.resizable(False, False)

        apply_dark_theme(self)

        ensure_default_users()
        self.config_data = load_config()

        ok, msg = check_or_init_license(self.config_data.get("months_valid", 12))
        if not ok:
            messagebox.showerror("License", msg)
            self.destroy()
            return

        # Frames
        self.login_frame = tk.Frame(self, padx=16, pady=16)
        self.main_frame  = tk.Frame(self, padx=16, pady=16)
        self.login_user  = None  # dict with username/role

        self.build_login()
        self.build_main()

        make_dark(self.login_frame)
        make_dark(self.main_frame)

        self.login_frame.pack(fill="both", expand=True)

    # ----- Login -----
    def build_login(self):
        f = self.login_frame
        tk.Label(f, text="MonaGraph®", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, columnspan=2, pady=(0,12))

        tk.Label(f, text="Username").grid(row=1, column=0, sticky="e", padx=(0,8), pady=4)
        self.ent_user = tk.Entry(f, width=28); self.ent_user.grid(row=1, column=1, pady=4, sticky="w")

        tk.Label(f, text="Password").grid(row=2, column=0, sticky="e", padx=(0,8), pady=4)
        self.ent_pass = tk.Entry(f, show="*", width=28); self.ent_pass.grid(row=2, column=1, pady=4, sticky="w")

        tk.Button(f, text="Login", width=18, command=self.on_login).grid(row=3, column=0, columnspan=2, pady=(12,6))

        self.lbl_info = tk.Label(f, text=f"Workbook: {self.config_data.get('workbook_file','(set in config.json)')}")
        self.lbl_info.grid(row=4, column=0, columnspan=2, pady=(8,0))

    def on_login(self):
        user = verify_login(self.ent_user.get().strip(), self.ent_pass.get().strip())
        if not user:
            messagebox.showerror("Login", "Invalid username or password.")
            return
        ok, msg = check_or_init_license(self.config_data.get("months_valid", 12))
        if not ok:
            messagebox.showerror("License", msg)
            return

        self.login_user = user
        self.login_frame.pack_forget()
        self.main_frame.pack(fill="both", expand=True)

        if user["role"] != "admin":
            self.btn_renew_license.config(state="disabled")

        self.refresh_sheets()

    # ----- Main -----
    def build_main(self):
        mf = self.main_frame

        header = tk.Frame(mf); header.pack(fill="x", pady=(0,10))
        tk.Label(header, text="MonaGraph", font=("Segoe UI", 20, "bold")).pack(side="left")
        self.btn_renew_license = tk.Button(header, text="Admin: Renew License", command=self.on_renew)
        self.btn_renew_license.pack(side="right")

        # Workbook row
        wb_row = tk.Frame(mf); wb_row.pack(fill="x", pady=(0,8))
        tk.Label(wb_row, text="Workbook:").pack(side="left")
        self.ent_wb = tk.Entry(wb_row, width=50); self.ent_wb.pack(side="left", padx=6)
        self.ent_wb.insert(0, self.config_data.get("workbook_file","TDM Charts.xlsm"))
        tk.Button(wb_row, text="Browse...", command=self.browse_wb).pack(side="left")

        # Patient sheet row
        ps_row = tk.Frame(mf); ps_row.pack(fill="x", pady=(0,12))
        tk.Label(ps_row, text="Patient Sheet:").pack(side="left")
        self.cbo_sheet = ttk.Combobox(ps_row, state="readonly", width=20, values=[])
        self.cbo_sheet.pack(side="left", padx=6)
        tk.Button(ps_row, text="Reload Sheets", command=self.refresh_sheets).pack(side="left")

        # Patient details
        details = tk.LabelFrame(mf, text="Patient Details (B2:B6)")
        details.pack(fill="x", pady=(0,10))
        self.ent_name = tk.Entry(details, width=30)
        self.ent_sex  = ttk.Combobox(details, values=["Male","Female"], state="readonly", width=12)
        self.ent_geno = tk.Entry(details, width=20)
        self.ent_dob  = tk.Entry(details, width=20)
        self.ent_cont = tk.Entry(details, width=20)
        self.chk_force = tk.IntVar(value=0)

        tk.Label(details, text="Name").grid(row=0, column=0, sticky="e", padx=6, pady=4); self.ent_name.grid(row=0, column=1, pady=4, sticky="w")
        tk.Label(details, text="Sex").grid(row=0, column=2, sticky="e", padx=6, pady=4); self.ent_sex.grid(row=0, column=3, pady=4, sticky="w")
        tk.Label(details, text="Genotype").grid(row=1, column=0, sticky="e", padx=6, pady=4); self.ent_geno.grid(row=1, column=1, pady=4, sticky="w")
        tk.Label(details, text="DOB").grid(row=1, column=2, sticky="e", padx=6, pady=4); self.ent_dob.grid(row=1, column=3, pady=4, sticky="w")
        tk.Label(details, text="Contact").grid(row=1, column=4, sticky="e", padx=6, pady=4); self.ent_cont.grid(row=1, column=5, pady=4, sticky="w")
        ttk.Checkbutton(details, text="Update details even if already set", variable=self.chk_force).grid(row=2, column=0, columnspan=6, sticky="w", padx=6, pady=(2,6))

        # ------- Variable selection + dynamic daily form -------
        sel_frame = tk.LabelFrame(mf, text="Select Variables to Record")
        sel_frame.pack(fill="x", pady=(0,6))

        # Internal keys -> (Label with units, default selected?)
        self.var_defs = {
            "date":          ("Date", True),
            "dose":          ("Dose (mg)", True),
            "blood_glucose": ("Blood Glucose (mg/dL)", False),
            "sbp":           ("Systolic BP (mm Hg)", False),
            "dbp":           ("Diastolic BP (mm Hg)", False),
            "weight":        ("Weight (kg)", False),
            "height":        ("Height (m)", False),
            "bmi":           ("BMI (kg/m²)", False),
            "inr":           ("INR", False),
            "lung_capacity": ("Lung Capacity (cm³)", False),
            "notes":         ("Notes", False),
        }

        self.var_selected = {}   # key -> BooleanVar
        col = 0; row = 0
        for key, (label, default_on) in self.var_defs.items():
            v = tk.BooleanVar(value=default_on)
            chk = ttk.Checkbutton(sel_frame, text=label, variable=v, command=self.update_daily_form)
            chk.grid(row=row, column=col, padx=6, pady=4, sticky="w")
            self.var_selected[key] = v
            col += 1
            if col % 3 == 0:
                col = 0
                row += 1

        self.rec_frame = tk.LabelFrame(mf, text="Daily Record Entries")
        self.rec_frame.pack(fill="x", pady=(0,10))

        # Storage for entry widgets and variables
        self.entry_vars = {}   # key -> tk.StringVar
        self.entries = {}      # key -> tk.Entry or tk.Text

        self.update_daily_form()

        # Buttons
        btns = tk.Frame(mf); btns.pack(fill="x", pady=(4,0))
        tk.Button(btns, text="Save to Excel", width=18, command=self.save_to_excel).pack(side="left", padx=(0,8))
        tk.Button(btns, text="Clear Daily Fields", command=self.clear_daily).pack(side="left")
        tk.Button(btns, text="Open Workbook", command=self.open_workbook).pack(side="right")

        # Status
        self.status = tk.Label(mf, text="", anchor="w")
        self.status.pack(fill="x", pady=(8,0))

    # Build / rebuild daily form based on checkboxes
    def update_daily_form(self):
        # clear current widgets
        for w in self.rec_frame.winfo_children():
            w.destroy()

        # create controls for selected vars
        r, c = 0, 0
        for key, selected in self.var_selected.items():
            if not selected.get():
                # if unchecked and previously had a value/widget, clean it
                if key in self.entry_vars: del self.entry_vars[key]
                if key in self.entries: del self.entries[key]
                continue

            label_text = self.var_defs[key][0]
            tk.Label(self.rec_frame, text=label_text).grid(row=r, column=c, sticky="e", padx=6, pady=4)

            sv = tk.StringVar()
            ent = None
            if key == "notes":
                ent = tk.Entry(self.rec_frame, width=40, textvariable=sv)
            elif key == "date":
                ent = tk.Entry(self.rec_frame, width=16, textvariable=sv)
            else:
                ent = tk.Entry(self.rec_frame, width=16, textvariable=sv)

            ent.grid(row=r, column=c+1, pady=4, sticky="w")
            self.entry_vars[key] = sv
            self.entries[key] = ent

            # default date to today for convenience
            if key == "date" and not sv.get():
                sv.set(datetime.date.today().isoformat())

            r += 1
            if r > 5:
                r = 0
                c += 2

        # Bind BMI auto-calc if weight/height present
        self.bind_bmi_autocalc()

        # re-apply dark colors to new widgets
        make_dark(self.rec_frame)

    def bind_bmi_autocalc(self):
        # When either weight or height changes, compute BMI = kg / (m^2)
        def calc_bmi(_evt=None):
            if "weight" in self.entry_vars and "height" in self.entry_vars and "bmi" in self.entry_vars:
                w = self.entry_vars["weight"].get().strip()
                h = self.entry_vars["height"].get().strip()
                try:
                    wv = float(w)
                    hv = float(h)
                    if hv > 0:
                        bmi = wv / (hv * hv)
                        self.entry_vars["bmi"].set(f"{bmi:.2f}")
                except Exception:
                    # ignore parse errors silently
                    pass

        # Unbind previous to avoid stacking
        for key in ("weight", "height"):
            if key in self.entries:
                try:
                    self.entries[key].unbind("<KeyRelease>")
                except Exception:
                    pass

        for key in ("weight", "height"):
            if key in self.entries:
                self.entries[key].bind("<KeyRelease>", calc_bmi)

    # ----- Actions -----
    def browse_wb(self):
        path = filedialog.askopenfilename(title="Select Excel Workbook",
                                          filetypes=[("Excel Macro-Enabled Workbook","*.xlsm"),
                                                     ("Excel Workbook","*.xlsx"),
                                                     ("All files","*.*")])
        if path:
            self.ent_wb.delete(0, tk.END)
            self.ent_wb.insert(0, path)
            self.status.config(text=f"Workbook set to: {path}")
            self.refresh_sheets()

    def refresh_sheets(self):
        wb_path = self.ent_wb.get().strip()
        if not wb_path or not os.path.exists(wb_path):
            self.cbo_sheet["values"] = []
            self.status.config(text="Workbook not found. Set the correct path.")
            return

        prev = self.cbo_sheet.get().strip()  # remember previous selection
        try:
            wb = load_workbook(wb_path, keep_vba=True, data_only=False)
            patients = get_patient_sheets(wb)
            wb.close()  # free file handle
            self.cbo_sheet["values"] = patients
            # CHANGED: keep previous selection if still present, else pick first
            if prev and prev in patients:
                self.cbo_sheet.set(prev)
            elif patients:
                self.cbo_sheet.set(patients[0])
            else:
                self.cbo_sheet.set("")
            self.status.config(text=f"Loaded sheets: {', '.join(patients) if patients else '(none)'}")
        except Exception as e:
            messagebox.showerror("Workbook Error", f"Failed to open workbook:\n{e}")

    def clear_daily(self):
        # clear all current entry vars (selected ones)
        for key, sv in list(self.entry_vars.items()):
            sv.set("")

        # also keep defaulting date if present
        if "date" in self.entry_vars:
            self.entry_vars["date"].set(datetime.date.today().isoformat())

    def open_workbook(self):
        path = self.ent_wb.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Open Workbook", "Workbook path is not set or file does not exist.")
            return
        try:
            if os.name == "nt":
                os.startfile(path)  # Windows
            else:
                messagebox.showinfo("Open Workbook", f"Open manually:\n{path}")
        except Exception as e:
            messagebox.showerror("Open Workbook", f"Could not open workbook:\n{e}")

    def save_to_excel(self):
        wb_path = self.ent_wb.get().strip()
        if not wb_path or not os.path.exists(wb_path):
            messagebox.showerror("Save", "Workbook path is not set or file does not exist.")
            return
        sheet_name = self.cbo_sheet.get().strip()
        if not sheet_name:
            messagebox.showerror("Save", "Please select a patient sheet.")
            return

        details = {
            "name": self.ent_name.get().strip(),
            "sex":  self.ent_sex.get().strip(),
            "genotype": self.ent_geno.get().strip(),
            "dob":  self.ent_dob.get().strip(),
            "contact": self.ent_cont.get().strip(),
        }

        # Build record only from selected/visible fields
        record = {}
        for key in self.entry_vars:
            record[key] = self.entry_vars[key].get().strip()

        force = bool(self.chk_force.get())

        try:
            wb = load_workbook(wb_path, keep_vba=True, data_only=False)
        except Exception as e:
            messagebox.showerror("Workbook Error", f"Failed to open workbook. Is it open in Excel?\n\n{e}")
            return

        if sheet_name not in wb.sheetnames:
            messagebox.showerror("Sheet", f"Sheet '{sheet_name}' not found.")
            wb.close()
            return

        ws = wb[sheet_name]

        try:
            write_patient_details(ws, details, force_update=force)
            row = append_daily_record(ws, record)
            wb.save(wb_path)
            wb.close()
            self.status.config(text=f"Saved to {sheet_name} at row {row}.")
            messagebox.showinfo("Saved", f"Data saved to '{sheet_name}' (row {row}). Open in Excel to refresh charts.")
            self.clear_daily()
        except PermissionError:
            wb.close()
            messagebox.showerror("Permission Error",
                                 "Could not save. Close the Excel file if it's open and try again.")
        except Exception as e:
            wb.close()
            messagebox.showerror("Save Error", f"An error occurred while saving:\n{e}")

    def on_renew(self):
        if not self.login_user or self.login_user.get("role") != "admin":
            messagebox.showwarning("Admin Only", "Only admin can renew the license.")
            return
        # Confirm admin password
        dlg = tk.Toplevel(self)
        dlg.title("Admin Password")
        make_dark(dlg)
        tk.Label(dlg, text="Re-enter admin password to renew").pack(padx=12, pady=(12,6))
        ent = tk.Entry(dlg, show="*"); ent.pack(padx=12, pady=6)
        def ok():
            if verify_login("admin", ent.get().strip()):
                admin_renew_license(self.config_data.get("months_valid", 12))
                dlg.destroy()
            else:
                messagebox.showerror("Auth", "Incorrect admin password.")
        tk.Button(dlg, text="Renew", command=ok).pack(pady=(4,12))

# ---------- Run ----------
if __name__ == "__main__":
    App().mainloop()
